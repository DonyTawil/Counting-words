#! python3
# A program that counts how many unique words exist in a pdf

import PyPDF2, os, logging
import argparse
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.stem import PorterStemmer, LancasterStemmer, WordNetLemmatizer
from openpyxl import load_workbook, Workbook


wordnet_lemmatizer = WordNetLemmatizer()
porter = PorterStemmer()
#lancaster = LancasterStemmer()

# Construct the argument parse and parse the arguments
ap = argparse.ArgumentParser()
ap.add_argument("-p", "--pdf", required = True, help = "Path to the pdf")
args = vars(ap.parse_args())

#logging.disable(logging.CRITICAL)
#To use logging: logging.debug(thing_to_output)
logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')

pdf = args["pdf"]    # Path to pdf

pdfFileObj = open(pdf, 'rb')
pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

all_words = []
lem_words = []
for page_num in range(pdfReader.numPages):
    pageObj = pdfReader.getPage(page_num)
    text = pageObj.extractText()
    text = text.lower()
    token_words = word_tokenize(text)
    for word in token_words:
         if word not in all_words:
             all_words.append(word)
         
         lem_word = wordnet_lemmatizer.lemmatize(word, pos="v")
         if lem_word not in lem_words:
            lem_words.append(lem_word)
             
print(len(all_words))
print(len(lem_words))


# Save the result to excel
dest_filename = 'count_words.xlsx'

try:
    wb = load_workbook(filename = dest_filename)
except FileNotFoundError:
    wb = Workbook()
    
ws = wb.create_sheet()

column_names = ["Unique words", "lemmas"]
words = [all_words, lem_words]

for i in range(len(column_names)):
    name = column_names[i]
    word_list = words[i]
    ws.cell(row=1, column=(i + 1), value=name)
    for j in range(len(word_list)):
        word = word_list[j]
        curr_row = j + 2
        ws.cell(row=curr_row, column=(i + 1), value=word)
        

wb.save(dest_filename)
